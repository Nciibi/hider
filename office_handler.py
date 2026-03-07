import os
from docx import Document
from openpyxl import load_workbook
from pptx import Presentation

class OfficeHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.ext = os.path.splitext(file_path)[1].lower()

    def get_metadata(self):
        """Returns core properties of the Office document."""
        if self.ext == '.docx':
            doc = Document(self.file_path)
            props = doc.core_properties
        elif self.ext == '.xlsx':
            wb = load_workbook(self.file_path)
            props = wb.properties
        elif self.ext == '.pptx':
            prs = Presentation(self.file_path)
            props = prs.core_properties
        else:
            raise ValueError(f"Unsupported office extension: {self.ext}")

        # Mapping: common_name -> (docx/pptx_attr, xlsx_attr)
        mapping = {
            "author": ("author", "creator"),
            "category": ("category", "category"),
            "comments": ("comments", "description"),
            "content_status": ("content_status", "contentStatus"),
            "created": ("created", "created"),
            "identifier": ("identifier", "identifier"),
            "keywords": ("keywords", "keywords"),
            "language": ("language", "language"),
            "last_modified_by": ("last_modified_by", "lastModifiedBy"),
            "last_printed": ("last_printed", "lastPrinted"),
            "modified": ("modified", "modified"),
            "revision": ("revision", "revision"),
            "subject": ("subject", "subject"),
            "title": ("title", "title"),
            "version": ("version", "version")
        }

        metadata = {}
        for common_name, attrs in mapping.items():
            attr = attrs[1] if self.ext == '.xlsx' else attrs[0]
            val = getattr(props, attr, None)
            metadata[common_name] = str(val) if val is not None else ""
        return metadata

    def update_metadata(self, key, value, output_path=None):
        """Updates a core property of the Office document."""
        if output_path is None:
            output_path = self.file_path

        # Mapping: common_name -> (docx/pptx_attr, xlsx_attr)
        mapping = {
            "author": ("author", "creator"),
            "comments": ("comments", "description"),
            "content_status": ("content_status", "contentStatus"),
            "last_modified_by": ("last_modified_by", "lastModifiedBy"),
            "last_printed": ("last_printed", "lastPrinted"),
        }
        
        real_key = key
        if self.ext == '.xlsx':
            real_key = mapping.get(key, (key, key))[1]
        else:
            real_key = mapping.get(key, (key, key))[0]

        if self.ext == '.docx':
            doc = Document(self.file_path)
            setattr(doc.core_properties, real_key, value)
            doc.save(output_path)
        elif self.ext == '.xlsx':
            wb = load_workbook(self.file_path)
            setattr(wb.properties, real_key, value)
            wb.save(output_path)
        elif self.ext == '.pptx':
            prs = Presentation(self.file_path)
            setattr(prs.core_properties, real_key, value)
            prs.save(output_path)
        else:
            raise ValueError(f"Unsupported office extension: {self.ext}")

        return output_path
